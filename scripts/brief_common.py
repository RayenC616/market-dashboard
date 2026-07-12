"""Shared logic for scripts/generate_daily_brief.py and generate_weekly_brief.py.

Both scripts: compute stats from data/market_history.xlsx, find S&P 500
best/worst movers over their own lookback window, ask Claude (with web
search) to research real news and write the brief narrative in Korean,
render into brief.html, and archive the outgoing brief under the right
category ("daily" or "weekly") in archive.html.
"""
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path

import anthropic
import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA_PATH = ROOT / "data" / "market_history.xlsx"
BRIEF_PATH = ROOT / "brief.html"
ARCHIVE_DIR = ROOT / "archive"
ARCHIVE_INDEX_PATH = ROOT / "archive.html"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

ASSET_META = {
    "SPX":    {"name": "S&P500 지수", "region": "미국", "kind": "index"},
    "NDX":    {"name": "나스닥종합지수", "region": "미국", "kind": "index"},
    "STOXX":  {"name": "STOXX 유럽600", "region": "유럽", "kind": "index"},
    "NKY":    {"name": "니케이225", "region": "일본", "kind": "index"},
    "KOSPI":  {"name": "코스피", "region": "한국", "kind": "index"},
    "HSI":    {"name": "항셍지수", "region": "홍콩", "kind": "index"},
    "WTI":    {"name": "WTI원유", "unit": "$/bbl", "kind": "commodity"},
    "GOLD":   {"name": "금", "unit": "$/oz", "kind": "commodity"},
    "SILVER": {"name": "은", "unit": "$/oz", "kind": "commodity"},
    "COPPER": {"name": "구리", "unit": "$/lb", "kind": "commodity"},
    "NICKEL": {"name": "니켈", "unit": "$/t", "kind": "commodity"},
    "CORN":   {"name": "옥수수", "unit": "¢/bu", "kind": "commodity"},
    "WHEAT":  {"name": "밀", "unit": "¢/bu", "kind": "commodity"},
    "SOY":    {"name": "대두", "unit": "¢/bu", "kind": "commodity"},
}
INDEX_ORDER = ["SPX", "NDX", "STOXX", "NKY", "KOSPI", "HSI"]
COMMODITY_ORDER = ["WTI", "GOLD", "SILVER", "COPPER", "NICKEL", "CORN", "WHEAT", "SOY"]


def last_non_null(values, from_idx):
    for i in range(from_idx, -1, -1):
        if values[i] is not None:
            return i
    return -1


def pct(cur, base):
    if base is None or base == 0:
        return None
    return (cur - base) / base * 100


def load_all_stats():
    wb = load_workbook(DATA_PATH, data_only=True)
    sheet = wb["Prices"]
    header = [c.value for c in sheet[1]]
    date_col = header.index("Date")

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    dates = [r[date_col] for r in rows]

    stats = {}
    series_by_asset = {}
    for key in ASSET_META:
        col = header.index(key)
        values = [r[col] for r in rows]
        series_by_asset[key] = values
        n = len(values)
        last_idx = last_non_null(values, n - 1)
        if last_idx < 0:
            continue
        last = values[last_idx]
        last_date = dates[last_idx]

        idx1 = last_non_null(values, last_idx - 1)
        idx5 = None
        count = 0
        for i in range(last_idx - 1, -1, -1):
            if values[i] is not None:
                count += 1
                if count == 5:
                    idx5 = i
                    break
        idx21 = None
        count = 0
        for i in range(last_idx - 1, -1, -1):
            if values[i] is not None:
                count += 1
                if count == 21:
                    idx21 = i
                    break

        year_start = datetime(last_date.year, 1, 1, tzinfo=timezone.utc)
        ytd_idx = None
        for i in range(last_idx, -1, -1):
            if dates[i].replace(tzinfo=timezone.utc) < year_start:
                ytd_idx = i
                break

        stats[key] = {
            "name": ASSET_META[key]["name"],
            "last": last,
            "last_date": last_date.date().isoformat(),
            "pct1D": pct(last, values[idx1]) if idx1 is not None and idx1 >= 0 else None,
            "pct1W": pct(last, values[idx5]) if idx5 is not None else None,
            "pct1M": pct(last, values[idx21]) if idx21 is not None else None,
            "ytd": pct(last, values[ytd_idx]) if ytd_idx is not None else None,
        }

    return stats, dates, series_by_asset


def fetch_sp500_tickers():
    resp = requests.get("https://en.wikipedia.org/wiki/List_of_S%26P_500_companies",
                         headers={"User-Agent": UA}, timeout=30)
    resp.raise_for_status()
    html = resp.text
    start = html.find('id="constituents"')
    table_html = html[start:start + 400000]
    end = table_html.find("</table>")
    table_html = table_html[:end]
    tickers = re.findall(r'<td id="[^"]*"><a[^>]*>([A-Z.]{1,6})</a></td>', table_html)
    return sorted(set(tickers))


def fetch_movers(tickers, lookback_days, top_n=3):
    """Best/worst S&P 500 movers over the trailing `lookback_days` trading
    days (1 for a daily brief, 5 for a weekly one)."""
    results = []
    for i in range(0, len(tickers), 20):
        batch = tickers[i:i + 20]
        try:
            resp = requests.get(
                "https://query1.finance.yahoo.com/v7/finance/spark",
                params={"symbols": ",".join(batch), "range": "1mo", "interval": "1d"},
                headers={"User-Agent": UA}, timeout=20,
            )
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            print(f"WARN: batch fetch failed ({batch[:3]}...): {e}")
            continue

        for item in data.get("spark", {}).get("result", []):
            symbol = item.get("symbol")
            try:
                response = item["response"][0]
                closes = [c for c in response["indicators"]["quote"][0]["close"] if c is not None]
                name = response["meta"].get("longName") or response["meta"].get("shortName") or symbol
                if len(closes) < lookback_days + 1:
                    continue
                last, base = closes[-1], closes[-(lookback_days + 1)]
                change = pct(last, base)
                if change is not None:
                    results.append({"symbol": symbol, "name": name, "last": last, "pct_change": change})
            except Exception:
                continue

    results.sort(key=lambda r: r["pct_change"], reverse=True)
    return results[:top_n], results[-top_n:][::-1]


def build_ytd_chart_data(dates, series_by_asset):
    year = dates[-1].year
    cutoff = datetime(year, 1, 1, tzinfo=timezone.utc)
    base_idx = {}
    for key in INDEX_ORDER:
        values = series_by_asset[key]
        for i in range(len(dates) - 1, -1, -1):
            if dates[i].replace(tzinfo=timezone.utc) < cutoff:
                base_idx[key] = i
                break

    master_dates = sorted({
        dates[i].date().isoformat()
        for key in INDEX_ORDER
        for i in range(base_idx.get(key, len(dates)), len(dates))
        if series_by_asset[key][i] is not None
    })
    date_index = {d: i for i, d in enumerate(dates_iso(dates))}

    series = {}
    for key in INDEX_ORDER:
        if key not in base_idx:
            continue
        base_val = series_by_asset[key][base_idx[key]]
        values = []
        for d in master_dates:
            idx = date_index.get(d)
            v = series_by_asset[key][idx] if idx is not None else None
            values.append(round(v / base_val * 100, 4) if v is not None else None)
        series[ASSET_META[key]["name"]] = values

    return {"dates": master_dates, "series": series}


def dates_iso(dates):
    return [d.date().isoformat() for d in dates]


BRIEF_TOOL = {
    "name": "submit_brief",
    "description": "Submit the completed market brief content in Korean.",
    "input_schema": {
        "type": "object",
        "properties": {
            "headline": {"type": "string", "description": "짧은 헤드라인 (아카이브 목록에 쓰일 한 문장 요약)"},
            "executive_summary": {"type": "string", "description": "5~8문장 분량의 요약 (한국어, 실제 수치와 근거를 포함)"},
            "drivers": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "title": {"type": "string"},
                        "analysis": {"type": "string"},
                    },
                    "required": ["title", "analysis"],
                },
                "minItems": 3, "maxItems": 4,
                "description": "주요 변동 요인 3~4개, 각각 제목과 분석 문단(한국어)",
            },
            "best_stocks_commentary": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {"symbol": {"type": "string"}, "why": {"type": "string"}},
                    "required": ["symbol", "why"],
                },
                "description": "제공된 최고 상승 종목 3개에 대한 한국어 코멘트",
            },
            "worst_stocks_commentary": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {"symbol": {"type": "string"}, "why": {"type": "string"}},
                    "required": ["symbol", "why"],
                },
                "description": "제공된 최저 상승(최대 하락) 종목 3개에 대한 한국어 코멘트",
            },
            "sources": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {"title": {"type": "string"}, "url": {"type": "string"}},
                    "required": ["title", "url"],
                },
                "description": "실제로 검색해 확인한 뉴스 출처 목록 (제목은 원문 언어 그대로)",
            },
        },
        "required": ["headline", "executive_summary", "drivers",
                     "best_stocks_commentary", "worst_stocks_commentary", "sources"],
    },
}


def call_claude(period_label, period_noun, movers_noun, stats, best, worst):
    """period_noun: '이번 주' or '오늘'. movers_noun: '주간' or '당일'."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError(
            "ANTHROPIC_API_KEY is not set. In GitHub: Settings -> Secrets and "
            "variables -> Actions -> New repository secret -> name it exactly "
            "ANTHROPIC_API_KEY and paste a valid key from console.anthropic.com. "
            "The workflow already forwards this secret into the environment "
            "(see .github/workflows/daily-brief.yml / weekly-brief.yml); if it's "
            "still missing here, the secret hasn't been created (or was created "
            "under a different name/repo) yet."
        )
    client = anthropic.Anthropic(api_key=api_key)

    index_lines = "\n".join(
        f"- {ASSET_META[k]['name']}: 종가 {stats[k]['last']:.2f}, 1일 {fmt_pct(stats[k]['pct1D'])}, "
        f"1주 {fmt_pct(stats[k]['pct1W'])}, 1개월 {fmt_pct(stats[k]['pct1M'])}, 연초대비 {fmt_pct(stats[k]['ytd'])}"
        for k in INDEX_ORDER if k in stats
    )
    commodity_lines = "\n".join(
        f"- {ASSET_META[k]['name']}: 종가 {stats[k]['last']:.4f} {ASSET_META[k]['unit']}, "
        f"1일 {fmt_pct(stats[k]['pct1D'])}, 1주 {fmt_pct(stats[k]['pct1W'])}"
        for k in COMMODITY_ORDER if k in stats
    )
    best_lines = "\n".join(f"- {s['symbol']} ({s['name']}): {movers_noun} {fmt_pct(s['pct_change'])}, 현재가 {s['last']:.2f}" for s in best)
    worst_lines = "\n".join(f"- {s['symbol']} ({s['name']}): {movers_noun} {fmt_pct(s['pct_change'])}, 현재가 {s['last']:.2f}" for s in worst)

    prompt = f"""당신은 기관 투자자용 시황 브리프를 작성하는 애널리스트입니다.
대상 기간: {period_label}

아래는 실제로 산출된 수치입니다 (반드시 이 수치를 그대로 인용하세요. 직접 계산하지 마세요):

[주요 지수]
{index_lines}

[원자재]
{commodity_lines}

[S&P500 {movers_noun} 최고 상승 종목 3개]
{best_lines}

[S&P500 {movers_noun} 최저 상승(최대 하락) 종목 3개]
{worst_lines}

작업 지시:
1. 웹 검색 도구를 사용해 {period_noun}({period_label}) 실제 시장 뉴스(반도체, 금리, 지정학, 원자재 등)를 조사하세요.
2. 조사한 내용과 위 수치를 근거로 submit_brief 도구를 호출해 다음을 한국어로 작성하세요:
   - headline: 아카이브 목록용 한 줄 요약
   - executive_summary: 5~8문장, 위 수치를 인용하며 {period_noun} 시장을 종합 요약
   - drivers: {period_noun} 시장을 움직인 핵심 요인 3~4개, 각각 제목과 분석(왜 시장에 영향을 미쳤는지)
   - best_stocks_commentary / worst_stocks_commentary: 제공된 종목별로 실제 검색된 근거를 바탕으로 1~2문장 코멘트
   - sources: 실제로 검색해서 확인한 기사 URL과 제목 (지어내지 마세요)
3. 반드시 submit_brief 도구 호출로 최종 결과를 제출하세요."""

    messages = [{"role": "user", "content": prompt}]
    # web_search is a server-side tool: Anthropic's API executes the search itself and
    # returns the results as extra content blocks in the same turn. It is NOT a
    # client-side tool, so we must never synthesize a "tool_result" for it ourselves —
    # only for genuinely client-side tools (here, none besides the terminal submit tool).
    # Verify "web_search_20250305" is still the current tool-type string in Anthropic's
    # docs before relying on this in production; the identifier may have moved on.
    tools = [{"type": "web_search_20250305", "name": "web_search"}, BRIEF_TOOL]

    for _ in range(6):
        response = client.messages.create(
            model="claude-sonnet-5",
            max_tokens=8000,
            tools=tools,
            messages=messages,
        )
        messages.append({"role": "assistant", "content": response.content})

        submit_call = next((b for b in response.content if b.type == "tool_use" and b.name == "submit_brief"), None)
        if submit_call:
            return submit_call.input

        if response.stop_reason in ("end_turn", "max_tokens"):
            messages.append({"role": "user", "content": "submit_brief 도구를 호출해 최종 브리프를 제출하세요."})
            continue

    raise RuntimeError("Claude did not submit a brief via submit_brief within the turn limit.")


def esc(text):
    return (str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def fmt_pct(value):
    if value is None:
        return "N/A"
    sign = "+" if value >= 0 else ""
    return f"{sign}{value:.2f}%"


def cls(value):
    if value is None:
        return ""
    return "pos" if value >= 0 else "neg"


def render_index_table(stats):
    rows = []
    for k in INDEX_ORDER:
        if k not in stats:
            continue
        s = stats[k]
        rows.append(
            f'    <tr data-key="{k}"><td>{ASSET_META[k]["name"]}</td><td>{ASSET_META[k]["region"]}</td>'
            f'<td>{s["last"]:,.2f}</td>'
            f'<td class="{cls(s["pct1D"])}">{fmt_pct(s["pct1D"])}</td>'
            f'<td class="{cls(s["pct1W"])}">{fmt_pct(s["pct1W"])}</td>'
            f'<td class="{cls(s["pct1M"])}">{fmt_pct(s["pct1M"])}</td>'
            f'<td class="{cls(s["ytd"])}">{fmt_pct(s["ytd"])}</td></tr>'
        )
    return "\n".join(rows)


def render_commodity_table(stats):
    rows = []
    for k in COMMODITY_ORDER:
        if k not in stats:
            continue
        s = stats[k]
        rows.append(
            f'    <tr data-key="{k}"><td>{ASSET_META[k]["name"]}</td><td>{s["last"]:,.4f} {ASSET_META[k]["unit"]}</td>'
            f'<td class="{cls(s["pct1D"])}">{fmt_pct(s["pct1D"])}</td>'
            f'<td class="{cls(s["pct1W"])}">{fmt_pct(s["pct1W"])}</td></tr>'
        )
    return "\n".join(rows)


def render_drivers(drivers):
    blocks = []
    for i, d in enumerate(drivers, 1):
        blocks.append(f'  <div class="callout">\n    <b>{i}. {esc(d["title"])}</b> {esc(d["analysis"])}\n  </div>')
    return "\n".join(blocks)


def render_stock_cards(entries, commentary_by_symbol, sign_class):
    cards = []
    for e in entries:
        why = commentary_by_symbol.get(e["symbol"], "")
        cards.append(f'''      <div class="stock-card">
        <div class="stock-row"><span class="stock-ticker">{esc(e["symbol"])}</span><span class="stock-chg {sign_class}">{fmt_pct(e["pct_change"])}</span></div>
        <div class="stock-name">{esc(e["name"])} — ${e["last"]:.2f}</div>
        <div class="stock-why">{esc(why)}</div>
      </div>''')
    return "\n".join(cards)


def render_sources(sources):
    return "<br>\n  ".join(f'<a href="{s["url"]}" target="_blank">{esc(s["title"])}</a>' for s in sources)


def render_brief(template_name, category, period_label, session_date, prepared_date, stats, best, worst, llm, chart_data):
    best_comment = {b["symbol"]: b["why"] for b in llm["best_stocks_commentary"]}
    worst_comment = {w["symbol"]: w["why"] for w in llm["worst_stocks_commentary"]}

    template = (ROOT / "scripts" / template_name).read_text(encoding="utf-8")
    html = template
    headline_escaped = esc(llm["headline"]).replace('"', "&quot;")
    html = html.replace("{{CATEGORY}}", category)
    html = html.replace("{{WEEK_LABEL}}", period_label)
    html = html.replace("{{HEADLINE_ESCAPED}}", headline_escaped)
    html = html.replace("{{SESSION_DATE}}", session_date)
    html = html.replace("{{PREPARED_DATE}}", prepared_date)
    html = html.replace("{{EXEC_SUMMARY}}", esc(llm["executive_summary"]))
    html = html.replace("{{INDEX_ROWS}}", render_index_table(stats))
    html = html.replace("{{COMMODITY_ROWS}}", render_commodity_table(stats))
    html = html.replace("{{DRIVERS}}", render_drivers(llm["drivers"]))
    html = html.replace("{{BEST_CARDS}}", render_stock_cards(best, best_comment, "pos"))
    html = html.replace("{{WORST_CARDS}}", render_stock_cards(worst, worst_comment, "neg"))
    html = html.replace("{{SOURCES}}", render_sources(llm["sources"]))
    html = html.replace("{{YTD_DATA_JSON}}", json.dumps(chart_data, ensure_ascii=False))
    html = html.replace("{{FILENAME}}", f"global_markets_brief_{session_date}.html")
    return html


def archive_previous_brief(new_session_date):
    """Archives whatever is currently in brief.html, tagging the archive.html
    entry with the category ("daily"/"weekly") read from its own BRIEF_META
    comment — i.e. a brief always archives itself as whatever category it
    was generated as, regardless of which script runs next."""
    if not BRIEF_PATH.exists():
        return
    old_html = BRIEF_PATH.read_text(encoding="utf-8")
    old_html = old_html.replace("location.href='index.html'", "location.href='../index.html'")
    old_html = old_html.replace("location.href='archive.html'", "location.href='../archive.html'")
    old_html = old_html.replace('src="js/asset-order.js"', 'src="../js/asset-order.js"')
    old_html = old_html.replace('src="js/auth.js"', 'src="../js/auth.js"')

    meta_match = re.search(
        r'<!-- BRIEF_META headline="([^"]*)" date_label="([^"]*)" category="([^"]*)" -->', old_html)
    if meta_match:
        old_headline_text = meta_match.group(1).replace("&quot;", '"')
        old_headline_date = meta_match.group(2)
        old_category = meta_match.group(3)
    else:
        # Bootstrap case: archiving the hand-written brief.html that predates
        # the daily/weekly split. It was originally written as a single-day brief.
        old_headline_text = "이전 세션 브리프"
        old_headline_date = "이전 브리프"
        old_category = "daily"

    date_for_filename = re.search(r"(\d{4}-\d{2}-\d{2})", old_headline_date)
    archived_name = (f"market_brief_{date_for_filename.group(1)}.html" if date_for_filename
                      else f"market_brief_before_{new_session_date}.html")

    ARCHIVE_DIR.mkdir(exist_ok=True)
    (ARCHIVE_DIR / archived_name).write_text(old_html, encoding="utf-8")

    if ARCHIVE_INDEX_PATH.exists():
        archive_html = ARCHIVE_INDEX_PATH.read_text(encoding="utf-8")
        new_entry = f'''  <a class="entry" data-category="{old_category}" href="archive/{archived_name}">
    <div>
      <div class="date">{old_headline_date}</div>
      <div class="headline">{old_headline_text}</div>
    </div>
    <span class="arrow">→</span>
  </a>
'''
        marker = f'  <div id="{old_category}Entries">\n'
        insert_at = archive_html.find(marker)
        if insert_at >= 0:
            insert_at += len(marker)
            archive_html = archive_html[:insert_at] + new_entry + archive_html[insert_at:]
            ARCHIVE_INDEX_PATH.write_text(archive_html, encoding="utf-8")
