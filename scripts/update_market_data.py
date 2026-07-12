"""Daily market data updater, run by .github/workflows/update-market-data.yml.

Fetches the latest close for all 14 tracked assets and appends one new row
to data/market_history.xlsx, unless today's row is already present.

Equity indices come from Yahoo Finance. Commodities come from investing.com
to match the historical CSVs the dataset was seeded from (Yahoo's commodity
futures occasionally diverge from investing.com's series/contract-roll
convention, which would otherwise create a visible seam in the chart).
"""
import re
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

DATA_PATH = Path(__file__).resolve().parent.parent / "data" / "market_history.xlsx"

YAHOO_SYMBOLS = {
    "SPX": "^GSPC",
    "NDX": "^IXIC",
    "STOXX": "^STOXX",
    "NKY": "^N225",
    "HSI": "^HSI",
    "KOSPI": "^KS11",
}
INVESTING_SLUGS = {
    "WTI": "crude-oil",
    "GOLD": "gold",
    "SILVER": "silver",
    "COPPER": "copper",
    "NICKEL": "nickel",
    "CORN": "us-corn",
    "WHEAT": "us-wheat",
    "SOY": "us-soybeans",
}
# Fallback source when investing.com is unreachable (it periodically blocks
# GitHub Actions' shared runner IPs). No reliable Yahoo ticker exists for
# nickel, so it has no fallback and stays blank on a bad day rather than
# guessing. Fallback values are raw Yahoo futures closes with no basis
# adjustment, so they may sit slightly off investing.com's own series for
# that one day (small, usually <5%) rather than match it exactly.
YAHOO_FALLBACK_SYMBOLS = {
    "WTI": "CL=F",
    "GOLD": "GC=F",
    "SILVER": "SI=F",
    "COPPER": "HG=F",
    "CORN": "ZC=F",
    "WHEAT": "ZW=F",
    "SOY": "ZS=F",
}
COPPER_TONNE_TO_LB = 2204.62
COLUMNS = ["SPX", "NDX", "STOXX", "NKY", "HSI", "KOSPI", "WTI", "GOLD", "SILVER",
           "COPPER", "NICKEL", "CORN", "WHEAT", "SOY"]

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")


def fetch_yahoo_last_close(symbol):
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
    resp = requests.get(url, params={"range": "5d", "interval": "1d"},
                        headers={"User-Agent": UA}, timeout=20)
    resp.raise_for_status()
    result = resp.json()["chart"]["result"][0]
    timestamps = result["timestamp"]
    closes = result["indicators"]["quote"][0]["close"]
    for ts, close in zip(reversed(timestamps), reversed(closes)):
        if close is not None:
            date = datetime.fromtimestamp(ts, tz=timezone.utc).date()
            return date, close
    return None, None


def fetch_investing_price(slug):
    resp = requests.get(f"https://www.investing.com/commodities/{slug}",
                         headers={"User-Agent": UA}, timeout=20)
    resp.raise_for_status()
    match = re.search(r'"last":([0-9.]+),"changePcr":(-?[0-9.]+)', resp.text)
    if not match:
        raise RuntimeError(
            f"page loaded but price pattern not found (status {resp.status_code}, "
            f"{len(resp.text)} bytes) - likely a bot-check/interstitial page"
        )
    return float(match.group(1))


def main():
    wb = load_workbook(DATA_PATH)
    sheet = wb["Prices"]
    header = [cell.value for cell in sheet[1]]
    date_col = header.index("Date") + 1

    last_row = sheet.max_row
    last_date_cell = sheet.cell(row=last_row, column=date_col).value
    last_date_str = last_date_cell.date().isoformat() if hasattr(last_date_cell, "date") else last_date_cell

    values = {}
    latest_dates = []
    for key, symbol in YAHOO_SYMBOLS.items():
        try:
            date, close = fetch_yahoo_last_close(symbol)
            if date is not None:
                values[key] = close
                latest_dates.append(date)
        except Exception as e:
            print(f"WARN: failed to fetch {key} ({symbol}): {e}")

    if not latest_dates:
        print("No equity data fetched, aborting without changes.")
        return

    target_date = max(latest_dates)
    target_date_str = target_date.isoformat()

    if last_date_str == target_date_str:
        print(f"Row for {target_date_str} already present, nothing to do.")
        return

    for key, slug in INVESTING_SLUGS.items():
        try:
            price = fetch_investing_price(slug)
            if key == "COPPER" and price is not None:
                price = price / COPPER_TONNE_TO_LB
            values[key] = price
        except Exception as e:
            print(f"WARN: failed to fetch {key} (investing.com/{slug}): {e}")
            fallback_symbol = YAHOO_FALLBACK_SYMBOLS.get(key)
            if fallback_symbol:
                try:
                    _, price = fetch_yahoo_last_close(fallback_symbol)
                    values[key] = price
                    print(f"  -> used Yahoo fallback {fallback_symbol}: {price}")
                except Exception as e2:
                    print(f"  -> Yahoo fallback {fallback_symbol} also failed: {e2}")

    new_row = [None] * len(header)
    new_row[date_col - 1] = target_date
    for key in COLUMNS:
        col_idx = header.index(key) + 1
        new_row[col_idx - 1] = values.get(key)

    sheet.append(new_row)
    new_row_idx = sheet.max_row
    sheet.cell(row=new_row_idx, column=date_col).number_format = "yyyy-mm-dd"
    wb.save(DATA_PATH)
    print(f"Appended row for {target_date_str}: {new_row}")


if __name__ == "__main__":
    main()
